#!/usr/bin/env python3

import json
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.utils import datetime as openpyxl_datetime
from openpyxl.styles import numbers
from requests.packages.urllib3.exceptions import InsecureRequestWarning
from bs4 import BeautifulSoup
import warnings
import pickle
import os
import sys
import argparse
import datetime
import pprint

pp = pprint.PrettyPrinter(indent=4)

cards = {}
xactions = {}

def save_prev_file(name, ext):
    old = ("_old."+ext).join(name.rsplit("."+ext, 1))
    if os.path.isfile(name):
        if os.path.isfile(old):
            os.remove(old)
        os.rename(name, old)

def handle_buyme(id, code):
    s = requests.Session()
    bj = s.get(
        url='https://buyme.co.il/',
        headers = {'User-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.0.0 Safari/537.36'},
    )
    if bj.status_code != 200:
        print("Unexpected error while updating", id, code)
        sys.exit(1)
    id = str(id)
    bj = s.get(url='https://buyme.co.il/siteapi/voucherBalance',
        params={'serialNumber':id, 'expiryDate':code},
        headers = {'User-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/97.0.4692.99 Safari/537.36'},
        verify=False)
    if bj.status_code != 200:
        print("Unexpected error while updating", id, code)
        sys.exit(1)
    b = json.loads(bj.text)
    #pp.pprint (bj.text)
    if b["type"] == -999:
        print("This buy me card is for a specific gift and has no balance")
        cards[id] = 0
        return
    if b["voucher"]["used"] == 1:
        cards[id] = 0
    else:
        cards[id] = b['value']
    creation_date = 'T'.join(b['voucher']['crspackage']['created_at'].split(' ')) + ".000"
    xactions[(creation_date, id)] = {'name': b['title'], 'deposit': True, 'sum': float(b['originalValue'])}
    for field in b['realizations']:
        d = 'T'.join(field['date'].split(' ')) + ".000"
        xactions[(d, id)] = {'name': field['redeemer'], 'deposit': False, 'sum': float(field['amount'])}

def handle_ybitan(id, code):
    #s = requests.Session()
    #bj = s.get(
    #    url='https://tavplus.mltp.co.il/',
    #    headers = {
    #        'User-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.0.0 Safari/537.36',
    #        'Accept' : 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
    #        'Accept-Encoding': 'gzip, deflate, br',
    #        'Accept-Language': 'he-IL,he;q=0.9',
    #    },
    #)
    #for c in s.cookies:
    #    print(c.name, ":", c.value)
    s = requests
    cookies = {
    #    # FortigateServer
    #    #'FGTServer' : 'BCEEC54F8E668B2C9A9D58EDE9E2373CCC1A7BC98A87F2222E2B84FBDA4F2EEC998E408F4A7601BDFA',
    #    # google analytics
        'activechatyWidgets' : '0',
        'chatyWidget_0' : '[{"k":"v-widget","v":"' + datetime.datetime.now().isoformat(timespec="milliseconds") + '"},{"k":"v-Whatsapp","v":"' + datetime.datetime.now().isoformat(timespec="milliseconds") + '"}]',
        '_ga' : 'GA1.3.104630825.1658012134',
    #    # Google cloud load balancer
    #    'GCLB' : 'CIP92LOJjK3eaA',
    #    'rbzid':'A70zU7of0H1Kut0CrQ/BvS7P5/u6Lqvwu3rT1BLZHC3VXfXM+HiEqOt6QRGrOIHQoNLFa+Xmqy4aaFK1mY1zgHp6nzkdxcDY2e+RNdAU6/+JE+88jmVO0O72e0QOIJ6L9z6riwE/mwt3bh6qKnSF0WmIeFwjXXwCB1Q7L6Q2h+rJNwF0CXJv7tn5vUptcNhFifwi5mS+Tnj+nKSbT/bwkkb6VfFo3/oQjGtsQxK8QuY1FNv/6nYUBikMbrFd8oLC	',
    #    'rbzsessionid':'4ca21ca2e15157e0410265a81be2a1cd',
        '_gid':'GA1.3.1329674055.1658012134',
    }
    headers = {
        'User-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.0.0 Safari/537.36',
    #    'Accept' : 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
    #    'Accept-Encoding': 'gzip, deflate, br',
    #    'Accept-Language': 'he-IL,he;q=0.9',
        'origin' : 'https://multipass.co.il',
        'authority' : 'multipass.co.il',
        'referer': 'https://multipass.co.il/%D7%91%D7%A8%D7%95%D7%A8-%D7%99%D7%AA%D7%A8%D7%94/',
    }
    bj = s.post(url='https://multipass.co.il/wp-admin/admin-ajax.php',
                headers=headers,
                cookies=cookies,
                data={'action':'getbuget', 'newcardid':id}, verify=True)
    #with open("err.html", "w") as f:
    #    print(bj.text, file=f)
    b = json.loads(bj.text)
    if b['ResultMessage'] != '' and b['ResultId'] == 0:
        cards[id] = int(b['UpdatedBugdet']) / 100.0
    bj = requests.post(url='https://multipass.co.il/wp-admin/admin-ajax.php',
        data={'CardId':id, 'action':'get_table'},
        headers=headers,
        cookies=cookies,
        verify=False)
    b = json.loads(bj.text)
    for field in b['data']:
        d = field['date']
        dep = field['LoadActualSum'] != ''
        xactions[(d, id)] = {'name': field['SupplierName'], 'deposit':dep, 'sum': field['LoadActualSum'] if dep else field['ApprovedSum']}


#
# Handle 
#
#https://stackoverflow.com/questions/59045550/cant-parse-the-username-to-make-sure-im-logged-in-to-a-website/59196651#59196651
def handle_max(id, code):
    code = code.split(':')
    s = requests.Session()
    bj = s.get(
        url='https://online.max.co.il/anonymous/giftcard/transactions.aspx',
        headers = {
            'User-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/97.0.4692.99 Safari/537.36',
            'Accept' : 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
            'Accept-Encoding': 'gzip, deflate, br',
            'Accept-Language': 'he-IL,he;q=0.9',
        },
    )
    #extra_cookies = {
    #    #'ct1' : 'c=fd0545ec-7272-435b-a4e9-03f7e6451154&e=2/8/2024 6:47:49 PM',
    #    #'ctFingerPrint' : 'V2luZG93cyBDaHJvbWVmYWxzZWhlLUlMMjQ4MTYyNTYwLDE0NDAyNTYwLDE0MDBBc2lhL0plcnVzYWxlbXRydWV0cnVldHJ1ZWZhbHNldHJ1ZW5vdCBhdmFpbGFibGVXaW4zMjdkMWY0NmFhYzE0ZGJhZDg1MDFiYjYxY2EwOWZkNGViYWY4MGNhZjEzN2U2Y2U0MjdlNzNhNjQ4M2U4NDdkYTMwNzhiNzY3MjZlNjJhNzEzN2ZmZGFhMGYyODAwODdlM2U2MGQ2NDFjMGZkNzU2YjBmYTU0ZTM5Y2RlZDRiODZiOWU4YTYyMzg5NzVlY2E3MDk2Yjg3ZDIzYmJlZTZlYTE4MjdjOGYzN2QyOGViM2MzMmU5OTM1ZTA2MWRiYTVhOGZhbHNlZmFsc2VmYWxzZWZhbHNlMCxmYWxzZSxmYWxzZTE1Y2RlYTZjY2NlOTI2ZjkyNjkxMTE5YmU1N2ViZDk0NzFmNmE4ZGE5MDJhZjZmM2FiNmJkMzA1ZjIwMzIzNWUxMjQuMDQzNDc1Mjc1MTYwNzQ=',
    #}
    #for c in extra_cookies:
    #    s.cookies.set(c, extra_cookies[c], domain='max.co.il')
    pp.pprint(s.cookies)
    if bj.status_code != 200:
        print("Unexpected error while updating", id, code)
        sys.exit(1)
    bj = s.post(url='https://online.max.co.il/anonymous/giftcard/transactions.aspx', data={
        'RequestToken' : '/wEFJDNiMGVmNWMxLTY5ZGYtNDI4My1iMzEzLTcxYTM4MmQ2YjQwOQ==',
        'ctl00$PlaceHolderMain$GiftCardTransactions1$txtMasterRegularUser' : id[0:4],
        'ctl00$PlaceHolderMain$GiftCardTransactions1$txtRegularUser2' : id[4:8],
        'ctl00$PlaceHolderMain$GiftCardTransactions1$txtRegularUser3' : id[8:12],
        'ctl00$PlaceHolderMain$GiftCardTransactions1$txtRegularUser4' : id[12:],
        'ctl00$PlaceHolderMain$GiftCardTransactions1$ddlCardYear:' : code[1],
        'ctl00$PlaceHolderMain$GiftCardTransactions1$ddlCardMonth' : code[0],
        'ctl00$PlaceHolderMain$GiftCardTransactions1$txtCvv' : code[2],
        'ctl00$MobileAppShow$hdnShowIos' : 'true',
        'ctl00$MobileAppShow$hdnShowAndroid' : 'false',
        }, verify=False)
    pp.pprint(bj.status_code)
    pp.pprint(bj.cookies)
    soup = BeautifulSoup(bj.text, 'html.parser')
    #soup = BeautifulSoup(open("resp.html"), 'html.parser')
    balance = soup.find('span', id='PlaceHolderMain_GiftCardTransactions1_lblBalance')
    cards[id] = float(balance.contents[0])
    print(id, cards[id])
    transactions = soup.find('table', id='tblGiftCardTransactions').find('tbody')
    rows = transactions.find_all('tr')
    for r in rows[1:-1]:
        tds = r.find_all('td')
        #print(tds)
        d = tds[1].contents[0].strip()
        print(d)
        n = tds[2].find('span').contents[0].strip()
        print(n)
        b = float(tds[4].contents[0].strip())
        print(b)

    #print(soup.find('div', id='PlaceHolderMain_GiftCardTransactions1_balanceSection'))
    print()
    sys.exit(1)

def handle_tav_zahav_8(id, code):
    bj = requests.post(url='https://www.shufersal.co.il/myshufersal/api/CardBalanceApi/GetCardBalanceAndTransactions', data={'cardNumber':id+str(code)}, verify=False)
    b = json.loads(bj.text)
    if b['HasCard'] == True and b['InactiveCard'] == False:
        cards[id] = b['CurrentBalance']
    elif b['HasCard'] == False:
        return
    for field in b['Transactions']:
        d = field['DateObject']
        dep = field['ActivityType'] == 'deposit'
        am = field['Amount']
        xactions[(d, id)] = {'name': field['BusinessName'], 'deposit':dep, 'sum':str(am) if dep else str(-am)}

def detect_paytment_method(id, code):
    l1 = len(str(id))
    l2 = len(str(code))
    if (l1, l2) == (16,10):
        handle_buyme(id, code)
    elif (l1, l2) == (8, 4):
        handle_ybitan(id, code)
    elif (l1, l2) == (16, 3) and str(id[0]) == '8':
        handle_tav_zahav_8(str(id), code)
    elif (l1, l2) == (16, 11) and str(id[0]) == '7':
        handle_max(str(id), code)
    else:
        print("Unknown ID:", id)
        return False
    return True

parser = argparse.ArgumentParser(description="""
This utility parses an XLSX file with a list of Prepaid cards.
You can then get a report of all your purchases saved to an output XLSX file.
Only new cards added to the input XLSX file are read from the internet.
Currently supported:
Yenot Bitan
Tav Hazahav (non-7215)
Buy Me
TODO:
Isracard
Max
""")
#parser.add_argument("-v", "--verbose", help="increase output verbosity", action="store_true", type=int)
#parser.add_argument("-q", "--quiet", action="store_true", help="quiet mode - only warn if all dupes or if more than one orig")
parser.add_argument("-f", "--data_file", metavar="picklefile", type=str, nargs=1, help="name of pickle file to store program state", default='cards_data.pickle')
parser.add_argument("-d", "--delete", metavar="card-ID", type=str, nargs=1, help="ID of card to remove")
parser.add_argument("-i", "--input", metavar="input-cards.xls", type=str, help="Name of XLSX file containing cards")
parser.add_argument("-o", "--output", metavar="output-transactions.xls", type=str, help="Name of XLSX file containing all transactions")
args = parser.parse_args()

if False:
    import logging

    # These two lines enable debugging at httplib level (requests->urllib3->http.client)
    # You will see the REQUEST, including HEADERS and DATA, and RESPONSE with HEADERS but without DATA.
    # The only thing missing will be the response.body which is not logged.
    try:
        import http.client as http_client
    except ImportError:
        # Python 2
        import httplib as http_client
    http_client.HTTPConnection.debuglevel = 1

    # You must initialize logging, otherwise you'll not see debug output.
    logging.basicConfig()
    logging.getLogger().setLevel(logging.DEBUG)
    requests_log = logging.getLogger("requests.packages.urllib3")
    requests_log.setLevel(logging.DEBUG)
    requests_log.propagate = True

try:
    with open(args.data_file, "rb") as f:
        cards = pickle.load(f)
        xactions = pickle.load(f)
        xls_file = pickle.load(f)
        output_xls_file = pickle.load(f)
except:
    pass

if args.input:
    xls_file = args.input
if args.output:
    output_xls_file = args.output

if args.delete:
    id = args.delete[0]
    if cards.get(id, None):
        del cards[id]
        print("Deleting", id)
    else:
        print("{} already deleted in cards[]".format(id))

if not xls_file:
    print("No input file defined")
    sys.exit(1)
if not output_xls_file:
    print("No output file defined")
    sys.exit(1)

wb = load_workbook(filename = xls_file, data_only=True)
sheet = wb.worksheets[0]

with warnings.catch_warnings():
    warnings.filterwarnings("ignore", category=InsecureRequestWarning)
    for i in range(sheet.min_row, sheet.max_row+1):
        id = sheet.cell(row=i, column=1).value
        code = sheet.cell(row=i, column=2).value
        if type(code) is datetime.datetime:
            code = code.strftime("%Y-%m-%d")
        if id == args.delete:
            continue
        if cards.get(id, -1) == 0:
            continue
        if not detect_paytment_method(id, code):
            continue
        v =  cards.get(id, -1)
        if  v > 0:
            print("ID: {}-{} => {}".format(id, code,  cards[id]))
        elif v == 0:
            print('ID:', id)
        else:
            print("Unknown card:", id)

if args.delete:
    x2 = xactions.copy()
    for (d, id) in x2.keys():
        if id == args.delete[0]:
            print("Deleting", d, id)
            del xactions[(d, id)]
    x2 = None

save_prev_file(args.data_file, "pickle")
with open(args.data_file, "wb") as f:
    pickle.dump(cards, f)
    pickle.dump(xactions, f)
    pickle.dump(xls_file, f)
    pickle.dump(output_xls_file, f)
    print("Saved", len(xactions.keys()), "transactions")

wb = Workbook()
ws = wb.active
ws.title = "קניות"
ws.cell(row=1, column=1).value = "תאריך"
ws.cell(row=1, column=2).value = "שעה"
ws.cell(row=1, column=3).value = "מיקום"
ws.cell(row=1, column=4).value = "סכום"
ws.cell(row=1, column=5).value = "טעינה"
ws.cell(row=1, column=6).value = "כרטיס"
row = 2
for (d, id) in sorted(xactions.keys(), key=lambda item:item[0]):
    fields = xactions[(d, id)]
    dt = openpyxl_datetime.from_ISO8601(d)
    ws.cell(row=row, column=1).value = dt.date()
    ws.cell(row=row, column=2).value = dt.time()
    v = fields['sum']
    dep = fields['deposit']
    if not v or v == '':
        v = 0.0
        #print(d, id, fields)
    ws.cell(row=row, column=3).value = fields['name']
    ws.cell(row=row, column=4).value = float(v) if not dep else 0
    ws.cell(row=row, column=4).number_format = numbers.FORMAT_NUMBER_00
    ws.cell(row=row, column=5).value = float(v) if dep else 0
    ws.cell(row=row, column=5).number_format = numbers.FORMAT_NUMBER_00
    ws.cell(row=row, column=6).value = id
    row += 1
save_prev_file(output_xls_file, "xlsx")
wb.save(output_xls_file)

    

#print(sheet_ranges['D18'].value)
